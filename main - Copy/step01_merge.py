# step01_merge.py

import os
import duckdb
import pandas as pd
from openpyxl import load_workbook
# Make sure this file exists in your project directory
from column_mapping import COLUMN_MAPPING

def main(source_directory, db_path, excel_path):
    """
    Merges all Excel files from a source directory into a new DuckDB database
    and a corresponding Excel file at the specified output paths.

    Args:
        source_directory (str): The full path to the folder with uploaded Excel files.
        db_path (str): The full path where the output DuckDB file should be saved.
        excel_path (str): The full path where the output Excel file should be saved.
    """
    # 1. Validate that the source directory exists
    if not os.path.exists(source_directory):
        print(f"❌ Error: Source directory '{source_directory}' not found.")
        return

    # 2. Find all Excel files in the source directory
    try:
        excel_files = [f for f in os.listdir(source_directory) if f.endswith(".xlsx")]
        if not excel_files:
            print(f"⚠️ Warning: No Excel (.xlsx) files found in '{source_directory}'.")
            # In a web app context, we should create empty files to avoid 404 errors
            pd.DataFrame().to_excel(excel_path, index=False)
            duckdb.connect(db_path).close()
            return
    except Exception as e:
        print(f"❌ Error listing files in source directory: {e}")
        return

    # 3. Connect to the new database at the specified path
    con = duckdb.connect(db_path)
    con.execute("DROP TABLE IF EXISTS final_data")

    all_data = []

    # 4. Process each Excel file found
    for file in excel_files:
        file_path = os.path.join(source_directory, file)
        try:
            # Use data_only=True to get cell values instead of formulas
            wb = load_workbook(file_path, data_only=True)
            print(f"   - Processing file: {file}")

            for sheet_name in wb.sheetnames:
                # Read the sheet into a pandas DataFrame, treating all data as strings
                df = pd.read_excel(file_path, sheet_name=sheet_name, dtype=str)
                
                # Rename columns based on the mapping dictionary
                df.rename(columns=COLUMN_MAPPING, inplace=True, errors="ignore")
                
                # Add a column to track the source filename
                df["FILENAME"] = file
                all_data.append(df)

        except Exception as e:
            print(f"   - ⚠️ Could not process file '{file}'. Error: {e}")

    # 5. Combine and save the data
    if all_data:
        final_df = pd.concat(all_data, ignore_index=True)
        
        # Create a table in the DuckDB database from the DataFrame
        con.execute("CREATE TABLE final_data AS SELECT * FROM final_df")
        
        # Save the combined DataFrame to the specified Excel file path
        final_df.to_excel(excel_path, index=False)
        print(f"✅ STEP 1: Merged {len(all_data)} dataframes successfully.")
    else:
        # If no data could be read, create empty outputs to prevent downstream errors
        print("⚠️ Warning: No data was successfully merged. Creating empty output files.")
        pd.DataFrame().to_excel(excel_path, index=False)

    # 6. Close the database connection
    con.close()